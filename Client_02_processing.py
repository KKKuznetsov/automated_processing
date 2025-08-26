# -*- coding: utf-8 -*-
r"""
Client_02_processing.py — одна запись (TASK_ID) -> один итоговый файл.
Поддержка csv / xls / xlsx. Нормализация 'День' и маркетинга.
Итог сохраняется в CSV (utf-8-sig, разделитель ';').
"""

import os
import csv
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import chardet

# === Пути ===
REESTR_PATH = Path(r"C:\Users\user\Desktop\Python_scripts\automated_processing\Reestr\new_files_registry.csv")
HEADER_PATH = Path(r"C:\Users\user\Desktop\Python_scripts\automated_processing\report_header\report_header.xlsx")
OUTPUT_DIR  = Path(r"C:\Users\user\Desktop\Итоговые отчеты")

CLIENT_NAME = "Client_02"
TARGET_REPORT_TYPE = "Type1"

FIELD_MAP = {
    "филиал": "supplier_filial",
    "клиент": "client_inlaw",
    "регион": "client_region",
    "город": "client_city",
    "улица": "client_adress",
    "товар": "tms",
    "инн клиента": "client_inn",
    "uid товара": "tms_id_ish",
    "день": "period",
    "аптека.ру": "market",
    "продажи, шт.": "amount_type_1",

    # из реестра:
    "file_path": "filename_ish",
    "client_name": "report_provider_name",
    "data_provider": "Report_Provaider",
}

def load_header_columns() -> list[str]:
    df_header = pd.read_excel(HEADER_PATH)
    return list(df_header.columns)

def load_registry() -> pd.DataFrame:
    return pd.read_csv(REESTR_PATH, sep=";", encoding="utf-8-sig")

def detect_encoding(file_path: Path) -> str:
    with open(file_path, "rb") as f:
        raw = f.read(50000)
    enc = chardet.detect(raw).get("encoding") or "utf-8"
    return enc

def read_csv(file_path: Path) -> pd.DataFrame:
    enc = detect_encoding(file_path)
    try:
        df = pd.read_csv(file_path, sep=";", encoding=enc)
    except Exception:
        df = pd.read_csv(file_path, sep=",", encoding=enc)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

def read_xlsx(file_path: Path) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active if len(wb.sheetnames) == 1 else wb[wb.sheetnames[0]]
    df = pd.DataFrame(ws.values)
    return df

def read_xls(file_path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(file_path, engine="xlrd")
    except Exception as e:
        print(f"[WARN] Нужен пакет 'xlrd' для .xls: {e}")
        try:
            return pd.read_excel(file_path)
        except Exception as e2:
            print(f"[WARN] Не удалось прочитать .xls: {e2}")
            return pd.DataFrame()

def find_header_row(df: pd.DataFrame) -> int | None:
    keys = set(k.lower() for k in FIELD_MAP.keys())
    for i in range(min(10, len(df))):
        row_vals = [str(x).strip().lower() for x in df.iloc[i]]
        if any(v in keys for v in row_vals):
            return i
    return None

def _excel_serial_to_dt(x) -> datetime:
    return datetime(1899, 12, 30) + timedelta(days=int(float(x)))

def parse_period(val) -> str | None:
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, (int, float)):
        try:
            return _excel_serial_to_dt(val).strftime("%d.%m.%Y")
        except Exception:
            pass
    s = str(val).strip()
    try:
        dt = datetime.strptime(s, "%d.%m.%Y")
        return dt.strftime("%d.%m.%Y")
    except Exception:
        pass
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%d.%m.%Y")
    except Exception:
        pass
    return None

def market_from_flag(flag_val: str) -> str:
    s = str(flag_val).strip().lower()
    if "да" in s:
        return "Аптека.ру"
    if "нет" in s:
        return "Коммерция"
    return "Коммерция" if s == "" else s

def normalize_excel_table(df_raw: pd.DataFrame) -> pd.DataFrame | None:
    if df_raw.empty:
        return None
    hdr = find_header_row(df_raw)
    if hdr is None:
        return None
    cols = [str(c).strip().lower() for c in df_raw.iloc[hdr]]
    df = df_raw.iloc[hdr + 1:].reset_index(drop=True)
    df.columns = cols
    return df

def transform(df: pd.DataFrame, reg_row: pd.Series, header_cols: list[str]) -> pd.DataFrame | None:
    out = pd.DataFrame()
    for src, tgt in FIELD_MAP.items():
        if src in df.columns:
            out[tgt] = df[src]
        elif src == "file_path":
            out[tgt] = reg_row["file_path"]
        elif src == "client_name":
            out[tgt] = reg_row["client_name"]
        elif src == "data_provider":
            out[tgt] = reg_row["data_provider"]

    if "period" in out.columns:
        out["period"] = out["period"].apply(parse_period)

    if "аптека.ру" in df.columns:
        out["market"] = df["аптека.ру"].apply(market_from_flag)

    for col in header_cols:
        if col not in out.columns:
            out[col] = None
    out = out[header_cols]
    return out if not out.empty else None

def main():
    task_id_env = os.getenv("TASK_ID")
    if not task_id_env or not task_id_env.isdigit():
        print("[INFO] TASK_ID не передан оркестратором — нечего делать.")
        return
    task_id = int(task_id_env)

    registry = load_registry()
    row_sel = registry[registry["id"].astype(int) == task_id]
    if row_sel.empty:
        print(f"[INFO] В CSV нет строки с id={task_id} (реестр обновлён?).")
        return
    row = row_sel.iloc[0]

    if str(row["client_name"]) != CLIENT_NAME or str(row["report_type"]) != TARGET_REPORT_TYPE:
        print(f"[INFO] id={task_id} не относится к {CLIENT_NAME}/{TARGET_REPORT_TYPE}. Пропуск.")
        return

    src = Path(row["file_path"])
    if not src.exists():
        print(f"[WARN] Файл не найден: {src}")
        return

    suf = src.suffix.lower()
    if suf == ".csv":
        df = read_csv(src)
    elif suf == ".xlsx":
        df = normalize_excel_table(read_xlsx(src))
        if df is not None:
            df.columns = [str(c).strip().lower() for c in df.columns]
    elif suf == ".xls":
        df_raw = read_xls(src)
        if df_raw.empty:
            print(f"[WARN] Пустая таблица XLS: {src}")
            return
        try:
            df_raw.columns = [str(c).strip().lower() for c in df_raw.columns]
            df = df_raw
        except Exception:
            df = normalize_excel_table(df_raw)
    else:
        print(f"[WARN] Неподдерживаемый формат: {suf}")
        return

    if df is None or df.empty:
        print(f"[WARN] Не удалось определить шапку/таблица пуста: {src}")
        return

    header_cols = load_header_columns()
    out = transform(df, row, header_cols)
    if out is None:
        print(f"[WARN] Пустой результат преобразования для id={task_id}")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = src.stem
    out_path = OUTPUT_DIR / f"{CLIENT_NAME}_id{task_id}_{base}_{ts}.csv"
    out.to_csv(out_path, sep=";", index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)
    print(f"[OK] Сохранён файл: {out_path}")

if __name__ == "__main__":
    main()
