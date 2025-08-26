# -*- coding: utf-8 -*-
r"""
Client_01_processing.py — обрабатывает одну запись (TASK_ID) -> один итоговый файл.
Поддержка csv / xls / xlsx. Для Excel ищем шапку в первых строках.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import chardet

# === Пути ===
REESTR_PATH = Path(r"C:\Users\user\Desktop\Python_scripts\automated_processing\Reestr\new_files_registry.csv")
HEADER_PATH = Path(r"C:\Users\user\Desktop\Python_scripts\automated_processing\report_header\report_header.xlsx")
OUTPUT_DIR  = Path(r"C:\Users\user\Desktop\Итоговые отчеты")

CLIENT_NAME = "Client_01"
TARGET_REPORT_TYPE = "Type1"

# Маппинг исходный_столбец -> итоговый_столбец (из report_header)
FIELD_MAP = {
    "инн": "client_inn",
    "клиент": "client_inlaw",
    "область_район": "client_region",
    "город": "client_city",
    "адрес": "client_adress",
    "название": "tms",
    "количество": "amount_type_1",
    "филиал": "supplier_filial",
    "код_клиента": "client_id_ish",
    "номер_документа": "naklad",
    "дата_документа": "naklad",
    "код_товара": "tms_id_ish",
    # из реестра:
    "file_path": "filename_ish",
    "client_name": "report_provider_name",
    "дата_документа_period": "period",
    "data_provider": "Report_Provaider",
}

def load_header_columns() -> list[str]:
    df_header = pd.read_excel(HEADER_PATH)
    return list(df_header.columns)

def load_registry() -> pd.DataFrame:
    return pd.read_csv(REESTR_PATH, sep=";", encoding="utf-8-sig")

def detect_encoding(file_path: Path) -> str:
    with open(file_path, "rb") as f:
        raw = f.read(50000)
    enc = chardet.detect(raw).get("encoding") or "utf-8"
    return enc

def read_csv(file_path: Path) -> pd.DataFrame:
    enc = detect_encoding(file_path)
    # Сначала пробуем ; затем ,
    try:
        df = pd.read_csv(file_path, sep=";", encoding=enc)
    except Exception:
        df = pd.read_csv(file_path, sep=",", encoding=enc)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

def read_xlsx(file_path: Path) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active if len(wb.sheetnames) == 1 else wb[wb.sheetnames[0]]
    df = pd.DataFrame(ws.values)
    return df

def read_xls(file_path: Path) -> pd.DataFrame:
    # Требуется xlrd для .xls
    try:
        return pd.read_excel(file_path, engine="xlrd")
    except Exception as e:
        print(f"[WARN] Нужен пакет 'xlrd' для чтения .xls: {e}")
        # fallback: попробуем стандартный движок (скорее всего не сработает для .xls)
        try:
            return pd.read_excel(file_path)
        except Exception as e2:
            print(f"[WARN] Не удалось прочитать .xls: {e2}")
            return pd.DataFrame()

def find_header_row(df: pd.DataFrame) -> int | None:
    keys = {k.lower() for k in FIELD_MAP.keys()}
    for i in range(min(10, len(df))):
        row_vals = [str(x).strip().lower() for x in df.iloc[i]]
        if any(v in keys for v in row_vals):
            return i
    return None

def parse_date(date_str):
    s = str(date_str).strip().replace("  ", " ")
    # Популярные форматы
    try:
        dt = datetime.strptime(s, "%d.%m.%Y")
        return dt.strftime("%d.%m.%Y")
    except Exception:
        pass
    try:
        dt = datetime.strptime(s, "%b %d %Y %I:%M%p")  # англ. месяцы
        return dt.strftime("%d.%m.%Y")
    except Exception:
        pass
    # RU месяцы в коротком виде
    months_ru = {'янв':1,'фев':2,'мар':3,'апр':4,'май':5,'июн':6,'июл':7,'авг':8,'сен':9,'окт':10,'ноя':11,'дек':12}
    parts = s.split()
    if len(parts) >= 3:
        mon = months_ru.get(parts[0].lower()[:3])
        try:
            day = int(parts[1]); year = int(parts[2])
            if mon:
                return datetime(year, mon, day).strftime("%d.%m.%Y")
        except Exception:
            return None
    return None

def normalize_excel_table(df_raw: pd.DataFrame) -> pd.DataFrame | None:
    if df_raw.empty:
        return None
    hdr = find_header_row(df_raw)
    if hdr is None:
        return None
    cols = [str(c).strip().lower() for c in df_raw.iloc[hdr]]
    df = df_raw.iloc[hdr + 1:].reset_index(drop=True)
    df.columns = cols
    return df

def transform(df: pd.DataFrame, reg_row: pd.Series, header_cols: list[str]) -> pd.DataFrame | None:
    out = pd.DataFrame()
    for src, target in FIELD_MAP.items():
        if src in df.columns:
            out[target] = df[src]
        elif src == "file_path":
            out[target] = reg_row["file_path"]
        elif src == "client_name":
            out[target] = reg_row["client_name"]
        elif src == "data_provider":
            out[target] = reg_row["data_provider"]
        elif src == "дата_документа_period" and "дата_документа" in df.columns:
            out[target] = df["дата_документа"].apply(parse_date)

    # Особая логика накладной
    if "номер_документа" in df.columns and "дата_документа" in df.columns:
        out["naklad"] = df["номер_документа"].astype(str) + " от " + df["дата_документа"].apply(lambda x: parse_date(x) or "")

    # Фильтр: количество != 0
    if "amount_type_1" in out.columns:
        out = out[out["amount_type_1"] != 0]

    # Добиваем до полного набора колонок из header + порядок
    for col in header_cols:
        if col not in out.columns:
            out[col] = None
    out = out[header_cols]
    return out if not out.empty else None

def main():
    # берём id задачи из окружения
    task_id_env = os.getenv("TASK_ID")
    if not task_id_env or not task_id_env.isdigit():
        print("[INFO] TASK_ID не передан оркестратором — нечего делать.")
        return
    task_id = int(task_id_env)

    registry = load_registry()
    row_sel = registry[registry["id"].astype(int) == task_id]
    if row_sel.empty:
        print(f"[INFO] В CSV нет строки с id={task_id} (реестр обновлён?).")
        return
    row = row_sel.iloc[0]

    # защита от «не своего» задания
    if str(row["client_name"]) != CLIENT_NAME or str(row["report_type"]) != TARGET_REPORT_TYPE:
        print(f"[INFO] id={task_id} не относится к {CLIENT_NAME}/{TARGET_REPORT_TYPE}. Пропуск.")
        return

    src_path = Path(row["file_path"])
    if not src_path.exists():
        print(f"[WARN] Файл не найден: {src_path}")
        return

    # читаем источник
    if src_path.suffix.lower() == ".csv":
        df = read_csv(src_path)
    elif src_path.suffix.lower() == ".xlsx":
        df_raw = read_xlsx(src_path)
        df = normalize_excel_table(df_raw)
        if df is not None:
            df.columns = [str(c).strip().lower() for c in df.columns]
    elif src_path.suffix.lower() == ".xls":
        df_raw = read_xls(src_path)
        if df_raw.empty:
            print(f"[WARN] Пустая таблица XLS: {src_path}")
            return
        # для .xls часто заголовки уже в первой строке, но приведём к нижнему регистру
        try:
            df_raw.columns = [str(c).strip().lower() for c in df_raw.columns]
            df = df_raw
        except Exception:
            df = normalize_excel_table(df_raw)
    else:
        print(f"[WARN] Неподдерживаемый формат: {src_path.suffix}")
        return

    if df is None or df.empty:
        print(f"[WARN] Не удалось определить шапку/таблица пуста: {src_path}")
        return

    header_cols = load_header_columns()
    out = transform(df, row, header_cols)
    if out is None:
        print(f"[WARN] Пустой результат преобразования для id={task_id}")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = src_path.stem
    out_path = OUTPUT_DIR / f"{CLIENT_NAME}_id{task_id}_{base}_{ts}.xlsx"
    out.to_excel(out_path, index=False)
    print(f"[OK] Сохранён файл: {out_path}")

if __name__ == "__main__":
    main()
